#!/usr/bin/env python3
################################################################################
### Excel BoM file generator
###
### This Scripts generates a BoM with KiBom, and then gets the output from that
### process and generates an Excel File with the contents
###
### Usage:
###     netlist: specifies the xml netlist to use to build the BoM
###     template: specifies the excel Template to use
###     --config: optional config file for KiBom, or KiBom_Config.ini in current dir by default
###     --variant: optional variant to use.
###
### Example:
### ExcelBomGenerator.py project.xml ExcelTemplate.xlsx
###
################################################################################

import argparse
import csv
import operator
import os
import re
import subprocess
import sys
from copy import copy, deepcopy

from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.drawing.image import Image

### Helper functions to sort columns "Naturally": C1, C2, C10,... and not C1, C10, C2,...
# returns a number as a int type if the text was a number, or the original value if it wasn't
def atoi(text):
    return int(text) if text.isdigit() else text
# The keyword function to get a natural sort on the Row. It selects the column to sort by
# (part reference), and then splits it text + digits. We sort on text firts, and then
# on number, but using the number as a number type, not a text type.
def natural_keys(row):
    text = row[0] # get the part reference for that column
    return [ atoi(c) for c in re.split('([0-9]+)',text) ]

# Just like range, for chars. Prints all chars from c1 to c2, inclusive
def char_range(c1, c2):
    """Generates the characters from `c1` to `c2`, inclusive."""
    for c in range(ord(c1), ord(c2)+1):
        yield chr(c)

# gets the style to apply to a column for the BoM table.
def getComponentStyle(rowNumber):
    # Get styles from legend
    if ws1["H8"].has_style:
        capStyle = copy(ws1["H8"]._style)
    if ws1["H9"].has_style:
        resStyle = copy(ws1["H9"]._style)
    if ws1["H10"].has_style:
        siliconStyle = copy(ws1["H10"]._style)
    if ws1["H11"].has_style:
        mechStyle = copy(ws1["H11"]._style)
    if ws1["H12"].has_style:
        inductorStyle = copy(ws1["H12"]._style)
    if ws1["H13"].has_style:
        cyrstalStyle = copy(ws1["H13"]._style)

    # dict with the component styles, based on their acronym
    componentDictionary = {
        "A": mechStyle,     # antenna
        "BT": mechStyle,    # Batteries
        "C": capStyle,      # caps
        "D": siliconStyle,  # diodes
        "LED": siliconStyle,# LEDs
        "G": mechStyle,     # logos
        "J": mechStyle,     # jumper / connector
        "JP": mechStyle,    # jumper / connector
        "L": inductorStyle, # inductors
        "Q": siliconStyle,  # transistors
        "R": resStyle,      # resistor
        "SW": siliconStyle, # switches
        "U": siliconStyle,  # Chips
        "X": siliconStyle,  # more chips
    }
    # Find key in the dictionary, return mechanical if it doesn't exist
    if ws1["B"+str(rowNumber)].value:
        match = re.match(r"([a-zA-Z]+)([0-9]+)", ws1["B"+str(rowNumber)].value, re.I)
        if match:
            key = match.groups()[0]
            return componentDictionary.get(key, mechStyle)
    return mechStyle # should be better to return "Normal" style, but this way we might spot errors

# Applyes the correct style to a row of the BoM table
def applyStyleToTableRow(rowNumber, tableWidth):

    componentStyle = getComponentStyle(rowNumber)
    # Table is from B to B + tableWidth
    for c in char_range('B',  chr(ord('B') + tableWidth - 1) ):
        ws1[c+str(rowNumber)]._style = copy(componentStyle)

# Returns true if the file with name passed as string exists, or false otherwise
def fileExists(filename):
    try:
        f = open(KibomConfig)
        f.close()
        return True
    except IOError:
        print('Config File is not accessible')
        return False

################################################################################
### Script Starts
################################################################################

### Parse Arguments ############################################################
parser = argparse.ArgumentParser()

# add a usage description
parser.usage = "Generates an Excel file with the Bill of Materials for the current project. The script needs an xml netlist from KiCad and an Excel File as template"
# add a positional arguments, they are mandatory and they are processed in order
# without any flags indicating what they are.
# The file to be processed
parser.add_argument("netlist", help="xml netlist file to process", type=str)
# Template to build the XLSX from
parser.add_argument("template", help="excel template to use as a base", type=str)
# Optional Config file
parser.add_argument("--config", help="Specifies a KiBom config file, or use 'KiBom_Config.ini' in the current folder as default", type=str)
# Variant specifier
parser.add_argument("--variant", help="variant of the board", type=str)
args = parser.parse_args()

### Run KiBoM to generate an CSV with the BoM ##################################
KibomConfig = "KiBom_config.ini"
if args.config != None:
    # Config file was provided as argument, replace default
    KibomConfig = args.config

# Check the provided files exist
if not fileExists(args.netlist):
    print("Netlist file " + args.netlist + " couldn't be found")
    exit(1)
if not fileExists(KibomConfig):
    print("Config file " + KibomConfig + " couldn't be found")
    exit(1)
if not fileExists(args.template):
    print("Template file " + args.template + " couldn't be found")
    exit(1)

command = ["/KiBoM/KiBOM_CLI.py", str(args.netlist), "out", "--cfg", str(KibomConfig), "-s", ";"]
if args.variant:
    command = command[:3] + ["--variant", args.variant] + command[3:]
KiBoMOutput = str(subprocess.check_output(command, shell=False))
# This generates "out_bom_.csv" if no variant is specified, or "out_bomm__(nameOfTheVariant).csv" if it is

### Open the output from KiBoM and rearrange the table #########################
projectName = os.path.splitext(args.netlist)[0]

# KiBoMOutput = "jasjajdj CSV Output -> test\n"

# Filenames
# match the output filename from KiBom with a regex
csv_in = ""
match = re.search(r'(.*)CSV Output -> (.*)\\n', KiBoMOutput)
if match:
    csv_in = match.group(2)
    print(csv_in)
else:
    print("No Matches! This is KiBom's output\n" + KiBoMOutput)
    sys.exit(1)

csv_out = "out_bom_v1.3_out.csv"
sortedCSV = csv_out

# Template to add the BoM to
BoMTemplate = args.template

print ("File name: "+ csv_in)
print ("Project name: "+ projectName)
try:
    with open(csv_in, newline='') as csvfile_in, open(csv_out, 'w+', newline='') as csvfile_out:
        reader = csv.reader(csvfile_in, delimiter=';', quotechar='|')
        writer = csv.writer(csvfile_out, delimiter=';', quotechar='|')
        newHeader = []
        extendedColumns = []
        extraRows = []
        for row in reader:
            # Get the part designator from the first row, and then split it in
            # letters and numbers
            if row:
                if row[0] == "References":
                    # First row, headers. Add two extra column headers
                    row.extend(["Sort1", "Sort2"])
                    newHeader = row
                # Check for the Letters-Numbers pattern in the reference
                else:
                    match = re.match(r"([a-zA-Z]+)([0-9]+)", row[0], re.I)
                    if match: # first column of row was part Reference
                        letters, numbers = match.groups()
                        #print(letters + "-" + numbers)
                        row.extend([letters, numbers])

                        extendedColumns.append(row)
                    else: # first column was not a part Reference
                        extraRows.append(row)

        sortedPartList = sorted(extendedColumns, key=natural_keys, reverse=False)

        # Reconstruct the output file
        writer.writerow(newHeader) # Write the new header
        writer.writerows(sortedPartList) # Write sorted component table
        writer.writerows([""]) # add a separator line before the misc info
        writer.writerows(extraRows) # misc info

        # close the files
        csvfile_out.close()
        csvfile_in.close()
except EnvironmentError: # parent of IOError, OSError *and* WindowsError where available
    print("Input File \"" + csv_in + "\" doesn't exist. Have you run KiBom?")
    exit()

### Create an Excel book from a template, and add the BoM table ################

### Open the excel spreadsheet and select the active sheet #####################
wb = Workbook()
wb = load_workbook(filename = BoMTemplate)

### Get the table of components from the CSV, and store the metadata in variables
partList = []
metadata = []
version = ""
totalComponents = ""
date = ""
variant = ""
with open(sortedCSV, newline='') as csvfile_in:
    reader = csv.reader(csvfile_in, delimiter=';', quotechar='|')
    for row in reader:
        # Get the part designator from the first row, and then split it in
        # letters and numbers
        if row:
            # Two types of non-empty rows, the table with the parts (len > 2) or
            # the metadata tail (len = 2)
            if len(row) == 2:
                if row[0] == "Schematic Version:":
                    version = row[1]
                elif row[0] == "Total components:":
                    totalComponents = row[1]
                elif row[0] == "Schematic Date:":
                    date = row[1]
                elif row[0] == "PCB Variant:":
                    variant = row[1].strip('\"')
            else:
                    # First row, headerm has References as first cell. Ignore if that's the case
                    # As long as we don't get an empty row, each row is part of the
                    # table of parts
                    partList.append(row)


### Start to work with the Excel Sheet. Get the active worksheet ###############
ws1 = wb.active
ws1.title = "Helix BoM"

### Store the BoM table in the excel worksheet #################################
# Set columns C (qty), and M (order2) as number
for col_cell in ws1['C']:
    col_cell.number_format = '0'
for col_cell in ws1['M']:
    col_cell.number_format = '0'

# Insert the Parts table
tableStartingRow = 15
tableStartingColumn = 2
noItemsInBOM = len(partList) -1 # -1 because of the header. We'll need this later
tableWidth = len(partList[0])

# Store the header
for colIndex, cellValue in enumerate(partList[0], start=tableStartingColumn):
    ws1.cell(column=colIndex, row=tableStartingRow, value=cellValue)

# Now store the parts
for rowIndex, row in enumerate(partList[1:], start=tableStartingRow+1):
    for colIndex, cellValue in enumerate(row, start=tableStartingColumn):
        # We know that all colums are text, except Qty, which is the second
        # column (Column C, index 3), and Order2, which is the last one
        if (colIndex != tableStartingColumn + 1) and (colIndex != tableWidth+tableStartingColumn):
            # Columns other than C and M, store as text
            ws1.cell(column=colIndex, row=rowIndex, value=cellValue)
        else:
            # Columns C and M, store as number
            ws1.cell(column=colIndex, row=rowIndex, value=int(cellValue))

### Make it a table. The template's table starts in B15, and ends in column M,
# (row 15 + # of items in BOM) #################################################

# quite convoluted way to calculate "table size is BX:MY"
tableRange = 'B'+ str(tableStartingRow) + ":" + chr(ord('B') + tableWidth - 1) + str((noItemsInBOM + tableStartingRow))
#print(tableRange)

tab = Table(displayName="BoM", ref=tableRange)
style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=False)
tab.tableStyleInfo = style
ws1.add_table(tab)

### Insert the Metadata ########################################################
# cell e8: date
if date == "":
    # If we didn't get the date from KiBom, use current time
    now = datetime.now() # current date and time
    date = now.strftime("%Y/%m/%d, %H:%M")
ws1.cell(column=5, row=8, value=date)
# cell e9: Title
if variant == "default":
    ws1.cell(column=5, row=9, value=projectName)
else:
    ws1.cell(column=5, row=9, value=projectName+ " - " + variant)
# cell e10: Revision
ws1.cell(column=5, row=10, value=version)
# cell e11: Company
# Uncomment if you want to add a company name
# ws1["E11"] = "Compuglobalhypermeganet"
# cell e12: Total Parts
# cell e13: unique parts
# E12 and 13 have formulas that need the title of the column C
columnTitle = ws1["C15"].value
ws1["E12"] = "=SUM(BoM[[#All],[" + columnTitle + "]])"
ws1["E12"].number_format = "0"
ws1["E13"] = "=COUNT(BoM[[#All],[" + columnTitle + "]])"
ws1["E13"].number_format = "0"

### Format Table ###############################################################
for row in range (tableStartingRow + 1, tableStartingRow + 1 + noItemsInBOM):
    applyStyleToTableRow(row, tableWidth)

### And done! ##################################################################
# Output filename
# File name is "Project_Name(_Variant)_vX.Y_BoM.xlsx"
dest_filename = ws1["E9"].value.replace(" ", "_")
if args.variant != None:
    dest_filename += "_" + ws1["E10"].value
dest_filename += '_BoM.xlsx'
print("File saved as " + dest_filename)
wb.save(filename = dest_filename)