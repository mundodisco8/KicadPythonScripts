#!/usr/bin/env python3
################################################################################
### Excel Positions file generator
###
### This script generates the Pick and Place positions for the board's components
### in Excel Format
### By default it creates a file named "*filename*_Positions.xlsx"
###
### Usage:
###
### positional arguments:
###   filename         pcb file to process
###
### optional arguments:
###   -h, --help       show this help message and exit
###   --output OUTPUT  name of the file with the componets' positions
###   --all            Prints the positions of all parts, including the ones
###
### Example:
### python3 PositionsGenerator.py PCB_TO_PROCESS.kicad_pcb
###
################################################################################

import os
import argparse
import re

from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

import pcbnew

# returns "TOP" if the component specified in component is on the top side of the
# board, and "BOTTOM" if it's on the bottom
def getTopOrBottom(component):
    if pcb.GetBoard().GetLayerName(component.GetLayer()) == "F.Cu":
        return "TOP"
    elif pcb.GetBoard().GetLayerName(component.GetLayer()) == "B.Cu":
        return "BOTTOM"
    else:
        return "WRONG LAYER!"

# Returns a string with the component type: Through Hole, SMD or Virtual
def getComponentTypeAsString(component):
    if component.GetAttributes() == 0:
        return "Through Hole"
    elif component.GetAttributes() == 1:
        return "SMD"
    elif component.GetAttributes() == 2:
        return "Virtual"
    else:
        return "UNKNOWN"

### Helper functions to sort columns "Naturally": C1, C2, C10,... and not C1, C10, C2,...
# returns a number as a int type if the text was a number, or the original value if it wasn't
def atoi(text):
    return int(text) if text.isdigit() else text
# The keyword function to get a natural sort on the Row. It selects the column to sort by
# (part reference), and then splits it text + digits. We sort on text firts, and then
# on number, but using the number as a number type, not a text type.
def natural_keys(row):
    text = row[0] # get the part reference for that column
    return [ atoi(c) for c in re.split('([0-9]+)',text) ]

def naturalSortTable(tableToSort):
        unsortedList = []
        header = []
        for rowIndex, row in enumerate(tableToSort):
            # Get the part designator from the first row, and then split it in
            # letters and numbers
            if row:
                if rowIndex == 0:
                    # First row, headers. Add two extra column headers
                    row.extend(["Sort1", "Sort2"])
                    header = row
                else:
                    # Check for the Letters-Numbers pattern in the reference
                    match = re.match(r"([a-zA-Z]+)([0-9]+)", row[0], re.I)
                    if match: # first column of row was part Reference
                        letters, numbers = match.groups()
                        #print(letters + "-" + numbers)
                        row.extend([letters, int(numbers)])
                        unsortedList.append(row)

        sortedList = sorted(unsortedList, key=natural_keys, reverse=False)
        sortedList.insert(0, header)
        return sortedList


### Argument Parser ############################################################
parser = argparse.ArgumentParser()

# add a usage description
parser.usage = "Generates a CSV file with the positions of the board components.\nBy default it creates a file named \"*filename* - Positions.csv\""
# add a positional arguments, they are mandatory and they are processed in order
# without any flags indicating what they are.
# The file to be processed
parser.add_argument("filename", help="pcb file to process", type=str)
# add an optional argument --output, in case we want to specify the CSV name
parser.add_argument("--output", help="CSV file with the componets' positions", type=str)
# add an optional argument --all to print also the position of virtual footprints
# action "store_true" sets it to true if it's specified, and false if it's not.
parser.add_argument("--all", help="Prints the positions of all parts, including the ones defined as virtual.", action="store_true")

args = parser.parse_args()


### Script starts here #########################################################

if os.path.exists(args.filename) and os.path.isfile(args.filename):
    pcb = pcbnew.LoadBoard(args.filename)
    projectName = os.path.splitext(args.filename)[0]
    print(projectName)
else:
    print("There's no file named '" + args.filename + "'!!")
    exit(1)

### Get board component's positions ############################################

boardComponents = pcb.GetBoard().GetModules()
# create the header fo the table based on the inclusion of --all
header = []
if args.all:
    header = ["Reference", "Footprint", "Value", "X Pos(mm)", "Y Pos(mm)", "Rotation", "PCB Side", "Type"]
else:
    header = ["Reference", "Footprint", "Value", "X Pos(mm)", "Y Pos(mm)", "Rotation", "PCB Side"]

positionsList = []
for component in boardComponents:
    c = component.GetCenter()
    # if --all is specified, print all the components positions
    if component.GetAttributes() == 2 and not args.all:
        # if component is virtual, and --all has NOT been specified, skipt this component
        continue
    # but if it isn't virtual, or it is, but --all has been specified, then add it to the list
    row = [component.GetReference(), component.GetFPID().GetLibItemName().c_str(), component.GetValue(), float(c.x / 10**6), float(c.y / 10 ** 6), float(component.GetOrientation() / 10), getTopOrBottom(component)]
    positionsList.append(row)

noOfRows = len(positionsList)
noOfColumns = len(positionsList[0])

# Add the header to the table, and then sort it naturally
positionsList.insert(0, header)
positionsList = naturalSortTable(positionsList)

### Open an Excel File #########################################################
wb = Workbook()
ws1 = wb.active
ws1.title = "Positions"

### Store the sorted table #####################################################
for rowIndex, row in enumerate(positionsList, start=1):
    for colIndex, cellValue in enumerate(row, start=1):
        if type(cellValue) == float:
            ws1.cell(column=colIndex, row=rowIndex, value=float(cellValue))
        if type(cellValue) == int:
            ws1.cell(column=colIndex, row=rowIndex, value=int(cellValue))
        else:
            ws1.cell(column=colIndex, row=rowIndex, value=cellValue)
noOfColumns = noOfColumns + 2 # we added sort1 and and sort2 to the table

### Create a table #############################################################
tableRange = "A1:" + get_column_letter(noOfColumns) + str(noOfRows + 1) # +1 because columns are 1-indexed in excel

tab = Table(displayName="Positions", ref=tableRange)
style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=False)
tab.tableStyleInfo = style
# find the column that has the TOP/BOTTOM values as they will determine the formatting
try:
    PCBSideColumnNo = positionsList[0].index("PCB Side") + 1
except:
    print("Couldn't find \"PCB Side\" in the header!")
    exit(1)

# Format rows blue if the component is on top, and green if it's on bottom
for rowIndex, row in enumerate(positionsList[1:], start=2):
    style = "Normal"
    if ws1.cell(column=PCBSideColumnNo, row=rowIndex).value == "TOP":
        style = "40 % - Accent1" # light blue
    else:
        style = "40 % - Accent3" # light green for bottom parts
    for colIndex, cellValue in enumerate(row, start=1):
        ws1.cell(column=colIndex, row=rowIndex).style = style

ws1.add_table(tab)

### Add a warning note #########################################################
# Add a note at the side about the origin of coordinates, and set style as Bad (red text over red BG)
# Add the notes a couple of columns after the table
noticeRowLetter = get_column_letter(noOfColumns + 2 + 2) # +2 of sort1 and sort2 + 2
ws1[noticeRowLetter+"2"].value = "All positions are relative to the Bottom-Left corner of the TOP Side of the PCB"
ws1[noticeRowLetter+"3"].value = "Positions of components in the Bottom side are given with board SEEN FROM TOP"
ws1[noticeRowLetter+"2"].style = "Bad"
ws1[noticeRowLetter+"3"].style = "Bad"
ws1.column_dimensions[noticeRowLetter].width = 100

### Save the file ##############################################################
# Substitute the spaces in projectName to underscores
projectName = projectName.replace(" ", "_")
print("File saved as " + projectName + "_Positions.xlsx")
wb.save(filename=projectName+"_Positions.xlsx")

